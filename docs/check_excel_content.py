#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
检查生成的Excel文件内容
"""

import pandas as pd
from openpyxl import load_workbook

def check_excel_content():
    """检查Excel文件内容"""
    excel_file = "/Users/gangwang/root/code/github/zan/zan-doc-agent/9种体质处方汇总.xlsx"
    
    try:
        # 使用openpyxl读取，保持格式
        wb = load_workbook(excel_file)
        ws = wb.active
        
        print("Excel文件内容预览:")
        print("=" * 80)
        
        # 读取前几行数据
        for row in range(1, min(12, ws.max_row + 1)):
            row_data = []
            for col in range(1, 5):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    # 截断长文本用于显示
                    if len(str(cell_value)) > 50:
                        cell_value = str(cell_value)[:50] + "..."
                    row_data.append(str(cell_value))
                else:
                    row_data.append("")
            
            print(f"第{row}行: {' | '.join(row_data)}")
        
        print("=" * 80)
        print(f"总行数: {ws.max_row}")
        print(f"总列数: {ws.max_column}")
        
    except Exception as e:
        print(f"读取Excel文件时出错: {e}")

if __name__ == "__main__":
    check_excel_content()
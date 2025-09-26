#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将9种体质结果页文案Word文档转换为Excel格式
按体质分类展示汤方/茶方、经络处方和中成药
"""

import os
import re
from docx import Document
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def read_word_document(file_path):
    """读取Word文档内容"""
    try:
        doc = Document(file_path)
        full_text = []
        
        for paragraph in doc.paragraphs:
            if paragraph.text.strip():
                full_text.append(paragraph.text.strip())
        
        return full_text
    except Exception as e:
        print(f"读取Word文档时出错: {e}")
        return []

def parse_constitution_data(text_lines):
    """解析体质数据 - 改进版本"""
    constitutions = {}
    
    # 9种体质名称及其变体
    constitution_patterns = {
        "平和质": ["平和质", "一、平和质"],
        "气虚质": ["气虚质", "八、气虚质"],
        "阳虚质": ["阳虚质", "五、阳虚质"],
        "阴虚质": ["阴虚质", "四、阴虚质"],
        "痰湿质": ["痰湿质", "二、痰湿质"],
        "湿热质": ["湿热质", "三、湿热质"],
        "血瘀质": ["血瘀质", "六、血瘀质"],
        "气郁质": ["气郁质", "七、气郁质"],
        "特禀质": ["特禀质", "六、特禀质"]
    }
    
    # 初始化数据结构
    for const_name in constitution_patterns.keys():
        constitutions[const_name] = {
            "汤方/茶方": [],
            "经络处方": [],
            "中成药": []
        }
    
    # 将所有文本合并为一个字符串，便于处理
    full_text = '\n'.join(text_lines)
    
    # 为每种体质提取信息
    for const_name, patterns in constitution_patterns.items():
        print(f"正在处理: {const_name}")
        
        # 找到体质开始位置
        start_pos = -1
        for pattern in patterns:
            pos = full_text.find(pattern)
            if pos != -1:
                start_pos = pos
                break
        
        if start_pos == -1:
            continue
        
        # 找到下一个体质的开始位置作为结束位置
        end_pos = len(full_text)
        for other_const, other_patterns in constitution_patterns.items():
            if other_const == const_name:
                continue
            for pattern in other_patterns:
                pos = full_text.find(pattern, start_pos + 1)
                if pos != -1 and pos < end_pos:
                    end_pos = pos
        
        # 提取该体质的文本段落
        const_text = full_text[start_pos:end_pos]
        
        # 解析汤方/茶方
        tang_fang_patterns = [
            r'【适宜吃的食物】(.*?)【不适宜吃的食物】',
            r'【推荐食谱】(.*?)(?=【|经络调理|中成药|$)',
            r'【茶饮推荐】(.*?)(?=【|经络调理|中成药|$)',
            r'【汤方推荐】(.*?)(?=【|经络调理|中成药|$)',
            r'【食疗建议】(.*?)(?=【|经络调理|中成药|$)',
            r'【饮食调理】(.*?)(?=【|经络调理|中成药|$)'
        ]
        
        for pattern in tang_fang_patterns:
            matches = re.findall(pattern, const_text, re.DOTALL | re.IGNORECASE)
            for match in matches:
                clean_text = re.sub(r'\s+', ' ', match.strip())
                if clean_text and len(clean_text) > 10:
                    constitutions[const_name]["汤方/茶方"].append(clean_text)
        
        # 解析经络处方
        jing_luo_patterns = [
            r'经络调理建议(.*?)(?=中成药建议|【|$)',
            r'【经络调理】(.*?)(?=【|中成药|$)',
            r'【穴位按摩】(.*?)(?=【|中成药|$)',
            r'【推拿按摩】(.*?)(?=【|中成药|$)'
        ]
        
        for pattern in jing_luo_patterns:
            matches = re.findall(pattern, const_text, re.DOTALL | re.IGNORECASE)
            for match in matches:
                clean_text = re.sub(r'\s+', ' ', match.strip())
                if clean_text and len(clean_text) > 5:
                    constitutions[const_name]["经络处方"].append(clean_text)
        
        # 解析中成药
        zhong_cheng_yao_patterns = [
            r'中成药建议(.*?)(?=【|经络调理|$)',
            r'【中成药推荐】(.*?)(?=【|经络调理|$)',
            r'【中成药】(.*?)(?=【|经络调理|$)'
        ]
        
        for pattern in zhong_cheng_yao_patterns:
            matches = re.findall(pattern, const_text, re.DOTALL | re.IGNORECASE)
            for match in matches:
                clean_text = re.sub(r'\s+', ' ', match.strip())
                if clean_text and len(clean_text) > 5:
                    constitutions[const_name]["中成药"].append(clean_text)
    
    return constitutions

def create_excel_file(constitutions_data, output_path):
    """创建Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "9种体质处方汇总"
    
    # 设置标题样式
    title_font = Font(name='微软雅黑', size=14, bold=True)
    header_font = Font(name='微软雅黑', size=12, bold=True)
    content_font = Font(name='微软雅黑', size=10)
    
    # 设置标题
    ws['A1'] = "9种体质处方汇总表"
    ws['A1'].font = title_font
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # 设置表头
    headers = ['体质类型', '汤方/茶方', '经络处方', '中成药']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # 填充数据
    row = 3
    constitution_order = ["平和质", "气虚质", "阳虚质", "阴虚质", "痰湿质", "湿热质", "血瘀质", "气郁质", "特禀质"]
    
    for constitution in constitution_order:
        if constitution in constitutions_data:
            data = constitutions_data[constitution]
            
            # 体质名称
            ws.cell(row=row, column=1, value=constitution).font = content_font
            
            # 汤方/茶方
            tang_fang = '\n\n'.join(data.get('汤方/茶方', []))
            if not tang_fang:
                tang_fang = "暂无相关信息"
            ws.cell(row=row, column=2, value=tang_fang).font = content_font
            
            # 经络处方
            jing_luo = '\n\n'.join(data.get('经络处方', []))
            if not jing_luo:
                jing_luo = "暂无相关信息"
            ws.cell(row=row, column=3, value=jing_luo).font = content_font
            
            # 中成药
            zhong_cheng_yao = '\n\n'.join(data.get('中成药', []))
            if not zhong_cheng_yao:
                zhong_cheng_yao = "暂无相关信息"
            ws.cell(row=row, column=4, value=zhong_cheng_yao).font = content_font
            
            row += 1
    
    # 调整列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 40
    
    # 设置行高和自动换行
    for row_num in range(2, row):
        ws.row_dimensions[row_num].height = 80
        for col in range(1, 5):
            ws.cell(row=row_num, column=col).alignment = Alignment(
                horizontal='left', 
                vertical='top', 
                wrap_text=True
            )
    
    # 保存文件
    wb.save(output_path)
    print(f"Excel文件已保存到: {output_path}")

def main():
    """主函数"""
    # 文件路径
    word_file = "/Users/gangwang/root/code/github/zan/zan-doc-agent/docs/9种体质结果页文案.docx"
    excel_file = "/Users/gangwang/root/code/github/zan/zan-doc-agent/9种体质处方汇总.xlsx"
    
    print("开始读取Word文档...")
    text_lines = read_word_document(word_file)
    
    if not text_lines:
        print("无法读取Word文档内容")
        return
    
    print(f"成功读取 {len(text_lines)} 行文本")
    
    print("开始解析体质数据...")
    constitutions_data = parse_constitution_data(text_lines)
    
    print(f"解析到 {len(constitutions_data)} 种体质数据:")
    for const_name, data in constitutions_data.items():
        print(f"  {const_name}: 汤方/茶方({len(data['汤方/茶方'])}), 经络处方({len(data['经络处方'])}), 中成药({len(data['中成药'])})")
    
    print("开始创建Excel文件...")
    create_excel_file(constitutions_data, excel_file)
    
    print("转换完成!")

if __name__ == "__main__":
    main()
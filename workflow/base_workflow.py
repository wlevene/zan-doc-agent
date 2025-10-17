import os
import json
import pandas as pd
from datetime import datetime
from typing import Dict, List, Any, Optional
from dataclasses import dataclass, asdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# 导入所需的代理类
from agents.wellness.wellness_mom_agent import WellnessMomAgent
from agents.scenario_generator.scenario_generator_agent import ScenarioGeneratorAgent
from agents.scenario_validator.scenario_validator_agent import ScenarioValidatorAgent
from agents.content_generator.content_generator_agent import ContentGeneratorAgent
from agents.content_validator.content_validator_agent import ContentValidatorAgent
from agents.product_recommender.product_recommender_agent import ProductRecommenderAgent
from agents.product_recommendation_validator.product_recommendation_validator_agent import ProductRecommendationValidatorAgent
from agents.content_rewriter.content_rewriter_agent import ContentRewriterAgent

@dataclass
class ContentItem:
    """文案数据项"""
    user_input: str = ""
    persona_detail: str = ""
    scenario_data: Dict[str, Any] = None
    scenario_validation_result: bool = False
    scenario_validation_reason: str = ""
    content_data: Dict[str, Any] = None
    content_validation_data: Dict[str, Any] = None
    content_validation_result: bool = False
    recommended_products: List[Dict[str, Any]] = None
    product_recommendation_reason: str = ""
    product_recommendation_success: bool = False
    product_recommendation_error: str = ""
    k3_code: str = ""  # K3编码
    product_name: str = ""  # 产品名称
    product_selling_points: str = ""  # 商品卖点
    formula_source: str = ""  # 配方出处
    product_price: str = ""  # 商品价格
    processing_stage: str = "pending"
    final_status: str = "pending"
    created_at: str = ""
    
    def __post_init__(self):
        if not self.created_at:
            self.created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if self.scenario_data is None:
            self.scenario_data = {}
        if self.content_data is None:
            self.content_data = {}
        if self.content_validation_data is None:
            self.content_validation_data = {}
        if self.recommended_products is None:
            self.recommended_products = []
    
    def to_dict(self) -> Dict[str, Any]:
        """转换为字典格式"""
        return asdict(self)
    
    def is_valid(self) -> bool:
        """检查是否为有效的文案数据（场景和内容都验收通过）"""
        return (self.scenario_validation_result and 
                self.content_validation_result and 
                self.final_status == "success")


class ContentCollector:
    """文案数据收集器"""
    
    def __init__(self, output_dir=None, k3_code=None):
        self.items: List[ContentItem] = []
        # 如果没有提供输出目录，则使用当前文件所在目录下的output目录
        if output_dir is None:
            output_dir = os.path.join(os.path.dirname(__file__), "output")
        self.output_dir = output_dir
        self.k3_code = k3_code  # 存储K3编码用于文件名生成
        os.makedirs(self.output_dir, exist_ok=True)
    
    def add_content(self, **kwargs) -> None:
        """添加内容项"""
        item = ContentItem(**kwargs)
        self.items.append(item)
    
    def add_scenario_only(self, user_input: str, persona_detail: str, 
                         scenario: str, scenario_validation_result: bool,
                         scenario_validation_reason: str) -> None:
        """仅添加场景数据（用于场景验收失败的情况）"""
        item = ContentItem(
            user_input=user_input,
            persona_detail=persona_detail,
            scenario_data={"content": scenario},
            scenario_validation_result=scenario_validation_result,
            scenario_validation_reason=scenario_validation_reason,
            processing_stage="scenario_only",
            final_status="failed" if not scenario_validation_result else "partial"
        )
        self.items.append(item)
    
    def get_valid_items(self) -> List[ContentItem]:
        """获取验收通过的文案数据"""
        return [item for item in self.items if item.is_valid()]
    
    def clear(self) -> None:
        """清空所有数据"""
        self.items.clear()
    
    def _get_original_content(self, item: ContentItem) -> str:
        """获取原始文案内容"""
        if not item.content_data:
            return ""
        
        # 如果明确标记为重写且有original_content字段，使用original_content
        if (item.content_data.get("rewritten", False) and 
            "original_content" in item.content_data):
            return item.content_data.get("original_content", "")
        
        # 否则使用content作为原始文案
        return item.content_data.get("content", "")
    
    def _clean_text_for_excel(self, text: str, preserve_newlines: bool = True) -> str:
        """清理文本，可选择是否保留换行符
        
        Args:
            text: 要清理的文本
            preserve_newlines: 是否保留换行符，默认为True
        """
        if not text:
            return ""
        
        # 转换为字符串（防止非字符串类型）
        text = str(text)
        
        if preserve_newlines:
            # 保留换行符，只清理其他问题字符
            text = text.replace('\r', '')  # 移除回车符（保留\n）
            text = text.replace('\t', ' ')  # 制表符替换为空格
            # 移除可能导致Excel解析问题的字符
            text = text.replace('"', "'")  # 双引号替换为单引号
            # 清理每行的多余空格，但保留换行结构
            lines = text.split('\n')
            cleaned_lines = [' '.join(line.split()) for line in lines]
            text = '\n'.join(cleaned_lines)
        else:
            # 原有逻辑：移除所有换行符
            text = text.replace('\n', ' ')  # 换行符替换为空格
            text = text.replace('\r', ' ')  # 回车符替换为空格
            text = text.replace('\t', ' ')  # 制表符替换为空格
            # 清理多余的空格
            text = ' '.join(text.split())
            # 移除可能导致Excel解析问题的字符
            text = text.replace('"', "'")  # 双引号替换为单引号
        
        return text
    
    def __len__(self) -> int:
        """返回数据总数"""
        return len(self.items)
    
    def get_count(self) -> int:
        """获取收集的数据总数（与 __len__ 功能相同）"""
        return len(self.items)
    
    def export_to_excel(self, filename: str = None, preserve_newlines: bool = True) -> Optional[str]:
        """导出数据到Excel文件，支持保留换行符
        
        Args:
            filename: Excel文件名，如果为None则自动生成
            preserve_newlines: 是否保留换行符，默认为True
        """
        if not self.items:
            return None
        
        if filename is None:
            if self.k3_code:
                # 优先使用K3编码作为文件名
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{self.k3_code}_{timestamp}.xlsx"
            else:
                # 如果没有K3编码，则使用原来的时间戳方式
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{timestamp}.xlsx"
        
        filepath = os.path.join(self.output_dir, filename)
        
        try:
            # 创建工作簿和工作表
            wb = Workbook()
            ws = wb.active
            ws.title = "文案数据"
            
            # 定义列标题和顺序
            headers = [
                "用户输入", "人设详情", "场景内容", "场景验收结果", "场景验收原因",
                "文案内容", "文案验收结果", "文案验收原因", "重写后文案",
                "K3编码", "产品名称", "产品描述", "产品价格",
                "商品推荐原因", "商品推荐成功", "商品推荐错误", "处理阶段", "最终状态", "创建时间"
            ]
            
            # 设置标题行样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # 写入标题行
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 定义内容样式
            content_font = Font(size=10)
            content_alignment = Alignment(
                horizontal="left", 
                vertical="top", 
                wrap_text=True  # 启用自动换行
            )
            
            # 写入数据行
            for row_idx, item in enumerate(self.items, 2):
                # 准备行数据
                row_data = self._prepare_row_data(item, preserve_newlines)
                
                # 写入每个单元格
                for col_idx, header in enumerate(headers, 1):
                    cell_value = row_data.get(header, "")
                    cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                    cell.font = content_font
                    cell.alignment = content_alignment
                    
                    # 对于包含换行符的单元格，调整行高
                    if preserve_newlines and cell_value and '\n' in str(cell_value):
                        line_count = str(cell_value).count('\n') + 1
                        # 根据行数调整行高（每行约15像素）
                        ws.row_dimensions[row_idx].height = max(20, line_count * 15)
            
            # 调整列宽
            self._adjust_column_widths(ws, headers)
            
            # 保存文件
            wb.save(filepath)
            return filepath
            
        except Exception as e:
            print(f"导出Excel失败: {str(e)}")
            return None
    
    def _prepare_row_data(self, item: ContentItem, preserve_newlines: bool = True) -> Dict[str, str]:
        """准备单行数据"""
        # 解析推荐商品信息
        product_name, product_description, product_price = self._parse_product_info(item, preserve_newlines)
        
        return {
            "用户输入": self._clean_text_for_excel(item.user_input, preserve_newlines),
            "人设详情": self._clean_text_for_excel(item.persona_detail, preserve_newlines),
            "场景内容": self._clean_text_for_excel(item.scenario_data.get("content", "") if item.scenario_data else "", preserve_newlines),
            "场景验收结果": "通过" if item.scenario_validation_result else "未通过",
            "场景验收原因": self._clean_text_for_excel(item.scenario_validation_reason, preserve_newlines),
            "文案内容": self._clean_text_for_excel(self._get_original_content(item), preserve_newlines),
            "文案验收结果": "通过" if item.content_validation_result else "未通过",
            "文案验收原因": self._clean_text_for_excel(item.content_validation_data.get("validation_reason", "") if item.content_validation_data else "", preserve_newlines),
            "重写后文案": self._clean_text_for_excel(item.content_data.get("content", "") if item.content_data else "", preserve_newlines),
            "K3编码": self._clean_text_for_excel(item.k3_code, preserve_newlines),
            "产品名称": product_name,
            "产品描述": product_description,
            "产品价格": product_price,
            "商品推荐原因": self._clean_text_for_excel(item.product_recommendation_reason, preserve_newlines),
            "商品推荐成功": "是" if item.product_recommendation_success else "否",
            "商品推荐错误": self._clean_text_for_excel(item.product_recommendation_error, preserve_newlines),
            "处理阶段": self._clean_text_for_excel(item.processing_stage, preserve_newlines),
            "最终状态": self._clean_text_for_excel(item.final_status, preserve_newlines),
            "创建时间": self._clean_text_for_excel(item.created_at, preserve_newlines)
        }
    
    def _parse_product_info(self, item: ContentItem, preserve_newlines: bool = True) -> tuple:
        """解析商品信息，返回(产品名称, 产品描述, 产品价格)"""
        product_name = ""
        product_description = ""
        product_price = ""
        
        # 优先使用ContentItem中已解析的product_name字段
        if item.product_name:
            product_name = self._clean_text_for_excel(item.product_name, preserve_newlines)
        
        # 如果product_name为空，则从recommended_products解析
        if not product_name and item.recommended_products:
            if isinstance(item.recommended_products, str):
                try:
                    product_data = json.loads(item.recommended_products)
                    if isinstance(product_data, dict):
                        product_name = self._clean_text_for_excel(str(product_data.get('name', '')), preserve_newlines)
                        product_description = self._clean_text_for_excel(str(product_data.get('description', product_data.get('desc', ''))), preserve_newlines)
                        product_price = self._clean_text_for_excel(str(product_data.get('price', '')), preserve_newlines)
                except (json.JSONDecodeError, AttributeError):
                    product_name = self._clean_text_for_excel(str(item.recommended_products), preserve_newlines)
            elif isinstance(item.recommended_products, dict):
                product_name = self._clean_text_for_excel(str(item.recommended_products.get('name', '')), preserve_newlines)
                product_description = self._clean_text_for_excel(str(item.recommended_products.get('description', item.recommended_products.get('desc', ''))), preserve_newlines)
                product_price = self._clean_text_for_excel(str(item.recommended_products.get('price', '')), preserve_newlines)
            elif isinstance(item.recommended_products, list) and len(item.recommended_products) > 0:
                first_product = item.recommended_products[0]
                if isinstance(first_product, dict):
                    product_name = self._clean_text_for_excel(str(first_product.get('name', '')), preserve_newlines)
                    product_description = self._clean_text_for_excel(str(first_product.get('description', first_product.get('desc', ''))), preserve_newlines)
                    product_price = self._clean_text_for_excel(str(first_product.get('price', '')), preserve_newlines)
        
        # 如果仍然为空，使用ContentItem中的其他字段
        if not product_name:
            product_name = self._clean_text_for_excel(item.product_name, preserve_newlines)
        if not product_description:
            product_description = self._clean_text_for_excel(item.product_selling_points, preserve_newlines)
        if not product_price:
            product_price = self._clean_text_for_excel(item.product_price, preserve_newlines)
        
        return product_name, product_description, product_price
    
    def _adjust_column_widths(self, ws, headers):
        """调整列宽"""
        column_widths = {
            "用户输入": 20,
            "人设详情": 30,
            "场景内容": 40,
            "场景验收结果": 12,
            "场景验收原因": 30,
            "文案内容": 50,
            "文案验收结果": 12,
            "文案验收原因": 30,
            "重写后文案": 50,
            "K3编码": 15,
            "产品名称": 25,
            "产品描述": 40,
            "产品价格": 12,
            "商品推荐原因": 30,
            "商品推荐成功": 12,
            "商品推荐错误": 30,
            "处理阶段": 15,
            "最终状态": 12,
            "创建时间": 20
        }
        
        for col_idx, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_idx)
            width = column_widths.get(header, 15)
            ws.column_dimensions[column_letter].width = width


class BaseWellnessWorkflow:
    """通用健康工作流基类"""
    
    def __init__(self, persona_detail: str, agent_config: Dict[str, Any] = None):
        self.persona_detail = persona_detail
        self.agent_config = agent_config or {}
        
        # 从agent_config中获取必要的参数
        wellness_config = self.agent_config.get("wellness_agent", {})
        api_key = wellness_config.get("api_key", "")
        base_url = wellness_config.get("base_url", "")
        app_id = wellness_config.get("app_id", "")
        
        # 初始化各个代理
        self.wellness_agent = WellnessMomAgent(api_key=api_key, base_url=base_url, app_id=app_id)
        self.scenario_generator = ScenarioGeneratorAgent()
        self.scenario_validator = ScenarioValidatorAgent()
        self.content_generator = ContentGeneratorAgent()
        self.content_validator = ContentValidatorAgent()
        # 初始化商品推荐相关代理
        self.product_recommender = ProductRecommenderAgent()
        prv_config = self.agent_config.get("product_recommendation_validator", {})
        prv_endpoint = prv_config.get("endpoint") or base_url or "http://119.45.130.88:8080/v1"
        prv_app_key = prv_config.get("app_key") or api_key or "app-oM9cjamwbeTy4em5KoEUvuDL"
        self.product_recommendation_validator = ProductRecommendationValidatorAgent(
            endpoint=prv_endpoint,
            app_key=prv_app_key
        )
        self.content_rewriter = ContentRewriterAgent()
    
    def run_complete_workflow(self, user_input: str, max_retries: int = 3) -> Dict[str, Any]:
        """运行完整的工作流，包括场景生成、验证、内容生成、验证、商品推荐和内容重写"""
        result = {
            "success": False,
            "data": None,
            "error": ""
        }
        
        try:
            # 1. 生成场景
            scenario = self._generate_scenario(user_input)
            if not scenario:
                result["error"] = "场景生成失败"
                return result
            
            # 2. 验证场景
            scenario_validation = self._validate_scenario(scenario, user_input)
            if not scenario_validation["valid"]:
                result["error"] = f"场景验证失败: {scenario_validation['reason']}"
                return result
            
            # 3. 生成内容
            content = self._generate_content(scenario, user_input)
            if not content:
                result["error"] = "内容生成失败"
                return result
            
            # 4. 验证内容（带重试机制）
            content_validation = self._validate_content_with_retry(content, scenario, user_input, max_retries)
            if not content_validation["valid"]:
                result["error"] = f"内容验证失败: {content_validation['reason']}"
                return result
            
            # 5. 推荐商品
            recommended_product = self._recommend_product(content_validation["content"], scenario)
            if not recommended_product["success"]:
                # 商品推荐失败，仍然返回已生成的内容
                result["success"] = True
                result["data"] = {
                    "scenario": scenario,
                    "content": content_validation["content"],
                    "recommended_product": None,
                    "product_recommendation_error": recommended_product["error"]
                }
                return result
            
            # 6. 验证商品推荐
            product_validation = self._validate_product_recommendation(
                recommended_product["product"], 
                content_validation["content"], 
                scenario
            )
            
            # 7. 重写内容以包含推荐商品
            rewritten_content = self._rewrite_content_with_product(
                content_validation["content"], 
                recommended_product["product"],
                scenario
            )
            
            # 8. 准备最终结果
            result["success"] = True
            result["data"] = {
                "scenario": scenario,
                "original_content": content_validation["content"],
                "rewritten_content": rewritten_content,
                "recommended_product": recommended_product["product"],
                "product_validation": product_validation
            }
            
        except Exception as e:
            result["error"] = f"工作流执行异常: {str(e)}"
        
        return result
    
    def _generate_scenario(self, user_input: str) -> str:
        """生成场景"""
        try:
            # 使用场景生成代理（process接口）创建场景
            resp = self.scenario_generator.process({
                "query": user_input,
                "persona": self.persona_detail,
                "persona_detail": self.persona_detail
            })
            if resp.success:
                return resp.content
            raise Exception(resp.error_message or "场景生成失败")
        except Exception as e:
            print(f"生成场景异常: {str(e)}")
            return None
    
    def _validate_scenario(self, scenario: str, user_input: str) -> Dict[str, Any]:
        """验证场景"""
        try:
            # 使用场景验证代理（process接口）验证场景
            resp = self.scenario_validator.process({
                "query": "请验收该场景是否合理且可执行",
                "scenario_to_validate": scenario,
                "user_input": user_input,
                "persona": self.persona_detail,
                "persona_detail": self.persona_detail
            })
            return {"valid": resp.success, "reason": resp.error_message or ""}
        except Exception as e:
            print(f"验证场景异常: {str(e)}")
            return {"valid": False, "reason": str(e)}
    
    def _generate_content(self, scenario: str, user_input: str) -> str:
        """生成内容"""
        try:
            # 使用内容生成代理（process接口）生成内容
            resp = self.content_generator.process({
                "query": user_input,
                "scenario_content": scenario,
                "persona": self.persona_detail,
                "persona_detail": self.persona_detail,
                "answer": ""
            })
            if resp.success:
                return resp.content
            raise Exception(resp.error_message or "内容生成失败")
        except Exception as e:
            print(f"生成内容异常: {str(e)}")
            return None
    
    def _validate_content_with_retry(self, content: str, scenario: str, user_input: str, max_retries: int) -> Dict[str, Any]:
        """验证内容（带重试机制）"""
        retry_count = 0
        current_content = content
        
        while retry_count < max_retries:
            try:
                # 将场景字符串截断到48字符以内以满足Dify输入限制
                scenario_for_validator = scenario[:48] if isinstance(scenario, str) else ""
                # 使用内容验收代理（process接口）验证内容
                resp = self.content_validator.process({
                    "query": "请验收以下文案是否符合场景、人设和规范",
                    "content_to_validate": current_content,
                    "scenario": scenario_for_validator,
                    "persona": self.persona_detail,
                    "persona_detail": self.persona_detail,
                    "answer": ""
                })
                if resp.success:
                    return {"valid": True, "content": current_content}
                
                # 如果验证失败，尝试使用重写代理（process接口）优化内容
                print(f"内容验收失败详情: error={resp.error_message}, content={resp.content}")
                reason = resp.content or resp.error_message or "文案不符合要求，请优化"
                rewrite_resp = self.content_rewriter.process({
                    "persona": self.persona_detail,
                    "scenario": scenario,
                    "query": current_content,
                    "feedback": reason
                })
                if rewrite_resp.success and rewrite_resp.content:
                    current_content = rewrite_resp.content
                
                retry_count += 1
            except Exception as e:
                print(f"验证或重写内容异常: {str(e)}")
                retry_count += 1
        
        # 达到最大重试次数仍然失败
        return {"valid": False, "reason": f"达到最大重试次数({max_retries})，内容仍未通过验证"}
    
    def _recommend_product(self, content: str, scenario: str) -> Dict[str, Any]:
        """推荐商品"""
        try:
            # 使用商品推荐代理（process接口）推荐商品
            resp = self.product_recommender.process({
                "query": "请根据内容与场景推荐一个最匹配的商品",
                "scenario": scenario,
                "user_profile": self.persona_detail,
                "persona": self.persona_detail,
                "inputs": {"content": content}
            })
            if resp.success:
                return {"success": True, "product": resp.content, "reason": ""}
            return {"success": False, "error": resp.error_message or "商品推荐失败"}
        except Exception as e:
            print(f"推荐商品异常: {str(e)}")
            return {"success": False, "error": str(e)}
    
    def _validate_product_recommendation(self, product: Dict[str, Any], content: str, scenario: str) -> Dict[str, Any]:
        """验证商品推荐"""
        try:
            # 使用商品推荐验收代理（process接口）验证推荐
            resp = self.product_recommendation_validator.process({
                "query": "请验收该商品推荐是否合理并符合要求",
                "recommendation_to_validate": product,
                "content": content,
                "scenario": scenario,
                "persona": self.persona_detail,
                "persona_detail": self.persona_detail
            })
            return {"valid": resp.success, "reason": resp.error_message or ""}
        except Exception as e:
            print(f"验证商品推荐异常: {str(e)}")
            return {"valid": False, "reason": str(e)}
    
    def _rewrite_content_with_product(self, content: str, product: Dict[str, Any], scenario: str) -> str:
        """重写内容以包含推荐商品"""
        try:
            # 使用内容重写代理（process接口）重写内容，加入商品信息
            resp = self.content_rewriter.process({
                "persona": self.persona_detail,
                "scenario": scenario,
                "query": content,
                "goods": product
            })
            return resp.content if resp.success and resp.content else content
        except Exception as e:
            print(f"重写内容异常: {str(e)}")
            # 如果重写失败，返回原始内容
            return content
#!/usr/bin/env python3
"""
Python script to process goods.xlsx file using typer CLI and pydantic models.
"""

import typer
from typer import style
from pathlib import Path
import shutil
from typing import Optional, List
from pydantic import BaseModel, Field, validator
from openpyxl import load_workbook
import sys
import re
from dify_client import DifyClient, WorkflowInput
from product_processor import DifyProductProcessor

app = typer.Typer()


class FileConfig(BaseModel):
    input_path: Path = Field(..., description="Path to the input Excel file")
    output_path: Optional[Path] = Field(None, description="Path for the output file")
    filter_keywords: List[str] = Field(
        ..., description="Keywords to filter out from product names"
    )
    product_name_column: str = Field(..., description="Column name for product names")

    @validator("input_path")
    def validate_input_path(cls, v):
        if not v.exists():
            raise ValueError(f"Input file does not exist: {v}")
        if v.suffix.lower() not in [".xlsx", ".xls"]:
            raise ValueError(f"Input file must be an Excel file (.xlsx or .xls): {v}")
        return v

    @validator("output_path", pre=True, always=True)
    def set_output_path(cls, v, values):
        if v is None and "input_path" in values:
            input_path = values["input_path"]
            stem = input_path.stem
            suffix = input_path.suffix
            output_path = input_path.parent / f"{stem}_processed{suffix}"
            return output_path
        return v


def copy_sheet_data(source_sheet, target_sheet):
    """Copy all data from source sheet to target sheet."""
    typer.echo(typer.style(f"Starting data copy from {source_sheet.title} to {target_sheet.title}", fg=typer.colors.BLUE))

    cell_count = 0
    formatted_count = 0

    for row in source_sheet.iter_rows():
        for cell in row:
            target_sheet[cell.coordinate] = cell.value
            cell_count += 1

            # Copy cell formatting
            if cell.has_style:
                target_cell = target_sheet[cell.coordinate]
                target_cell.font = cell.font
                target_cell.border = cell.border
                target_cell.fill = cell.fill
                target_cell.number_format = cell.number_format
                target_cell.protection = cell.protection
                target_cell.alignment = cell.alignment
                formatted_count += 1

    typer.echo(typer.style(f"Copied {cell_count} cells with {formatted_count} formatted cells", fg=typer.colors.GREEN))


def find_product_name_column(sheet, column_name: str) -> Optional[int]:
    """Find the column index for the product name column."""
    typer.echo(typer.style(f"Searching for column '{column_name}' in sheet '{sheet.title}'", fg=typer.colors.BLUE))
    header_row = sheet[1]  # Assume first row contains headers

    for idx, cell in enumerate(header_row, 1):
        if cell.value:
            cell_value = str(cell.value).strip()
            if cell_value == column_name:
                typer.echo(typer.style(f"Found target column '{column_name}' at index {idx}", fg=typer.colors.GREEN))
                return idx

    typer.echo(typer.style(f"Column '{column_name}' not found", fg=typer.colors.RED), err=True)
    return None


def clear_filtered_rows(
    sheet, product_name_col: int, filter_keywords: List[str]
) -> int:
    """Clear rows that contain any of the filter keywords in product name column."""
    typer.echo(typer.style(f"Starting row filtering with keywords: {filter_keywords}", fg=typer.colors.BLUE))
    cleared_count = 0
    max_col = sheet.max_column
    total_rows = sheet.max_row - 1  # Exclude header row

    # Iterate through all rows (skip header row)
    for row_num in range(2, sheet.max_row + 1):
        product_name_cell = sheet.cell(row=row_num, column=product_name_col)

        if product_name_cell.value:
            product_name = str(product_name_cell.value).strip()

            # Check if any filter keyword is in the product name
            matched_keywords = [
                keyword for keyword in filter_keywords if keyword in product_name
            ]

            if matched_keywords:
                # Clear all cells in this row but keep the row structure
                for col_num in range(1, max_col + 1):
                    cell = sheet.cell(row=row_num, column=col_num)
                    cell.value = None
                cleared_count += 1

    typer.echo(typer.style(f"Cleared {cleared_count} out of {total_rows} rows", fg=typer.colors.YELLOW))
    return cleared_count


def clean_product_name(product_name: str) -> str:
    """Remove content between 【 and 】 characters from product name."""
    if not product_name:
        return product_name

    # Remove content between 【 and 】 (including the brackets)
    cleaned_name = re.sub(r"【[^】]*】", "", product_name)

    # Clean up extra spaces
    cleaned_name = re.sub(r"\s+", " ", cleaned_name).strip()

    return cleaned_name


def clean_product_names(sheet, product_name_col: int) -> int:
    """Clean product names by removing content between 【 and 】 characters."""
    typer.echo(typer.style(f"Starting product name cleaning for column {product_name_col}", fg=typer.colors.BLUE))
    cleaned_count = 0
    total_rows = sheet.max_row - 1  # Exclude header row

    # Iterate through all rows (skip header row)
    for row_num in range(2, sheet.max_row + 1):
        product_name_cell = sheet.cell(row=row_num, column=product_name_col)

        if product_name_cell.value:
            original_name = str(product_name_cell.value).strip()
            cleaned_name = clean_product_name(original_name)

            if cleaned_name != original_name:
                product_name_cell.value = cleaned_name
                cleaned_count += 1

    typer.echo(typer.style(f"Cleaned {cleaned_count} out of {total_rows} product names", fg=typer.colors.YELLOW))
    return cleaned_count


def process_excel_file(config: FileConfig) -> None:
    """Process the Excel file according to requirements."""
    typer.echo(typer.style("=== Starting Excel file processing ===", fg=typer.colors.MAGENTA, bold=True))
    typer.echo(typer.style(f"Configuration: Input={config.input_path}, Output={config.output_path}", fg=typer.colors.CYAN))
    typer.echo(typer.style(f"Filter keywords: {config.filter_keywords}", fg=typer.colors.CYAN))
    typer.echo(typer.style(f"Product column: {config.product_name_column}", fg=typer.colors.CYAN))
    
    # 初始化 Dify 客户端
    dify_client = DifyClient(
        base_url="http://119.45.130.88:8080/v1",
        api_key="your_api_key_here"  # 请替换为实际的 API Key
    )
    typer.echo(typer.style("Dify client initialized", fg=typer.colors.GREEN))

    try:
        # Step 1: Copy the original file to a new file
        typer.echo(typer.style("Step 1: Copying original file", fg=typer.colors.BLUE, bold=True))
        shutil.copy2(config.input_path, config.output_path)
        typer.echo(typer.style(f"File copied successfully to {config.output_path}", fg=typer.colors.GREEN))

        # Step 2: Open the new file
        typer.echo(typer.style("Step 2: Loading Excel workbook", fg=typer.colors.BLUE, bold=True))
        workbook = load_workbook(config.output_path)
        typer.echo(typer.style(
            f"Workbook loaded with {len(workbook.sheetnames)} sheets: {workbook.sheetnames}", fg=typer.colors.GREEN
        ))

        # Step 3: Get the first sheet (sheet0)
        if len(workbook.sheetnames) == 0:
            typer.echo(typer.style("The Excel file has no sheets", fg=typer.colors.RED), err=True)
            raise ValueError("The Excel file has no sheets")

        original_sheet = workbook.worksheets[0]
        original_sheet_name = original_sheet.title

        typer.echo(typer.style(f"Step 3: Found original sheet '{original_sheet_name}'", fg=typer.colors.BLUE, bold=True))
        typer.echo(typer.style(
            f"Original sheet dimensions: {original_sheet.max_row} rows, {original_sheet.max_column} columns", fg=typer.colors.GREEN
        ))

        # Step 4: Create a new sheet named 'sheet1'
        new_sheet_name = "sheet1"
        typer.echo(typer.style(f"Step 4: Creating new sheet '{new_sheet_name}'", fg=typer.colors.BLUE, bold=True))
        new_sheet = workbook.create_sheet(title=new_sheet_name)
        typer.echo(typer.style(f"New sheet '{new_sheet_name}' created successfully", fg=typer.colors.GREEN))

        # Step 5: Copy data from original sheet to new sheet
        typer.echo(typer.style("Step 5: Copying data from original sheet to new sheet", fg=typer.colors.BLUE, bold=True))
        copy_sheet_data(original_sheet, new_sheet)
        typer.echo(typer.style("Data copy completed", fg=typer.colors.GREEN))

        # Step 6: Find the product name column
        typer.echo(typer.style(
            f"Step 6: Finding product name column '{config.product_name_column}'", fg=typer.colors.BLUE, bold=True
        ))
        product_name_col = find_product_name_column(
            new_sheet, config.product_name_column
        )

        if product_name_col is None:
            typer.echo(typer.style(
                f"Could not find column '{config.product_name_column}' in the sheet", fg=typer.colors.RED
            ), err=True)
            typer.echo(typer.style("Available columns in first row:", fg=typer.colors.YELLOW))
            for idx, cell in enumerate(new_sheet[1], 1):
                if cell.value:
                    typer.echo(typer.style(f"  Column {idx}: {cell.value}", fg=typer.colors.WHITE))
            typer.echo(typer.style("Skipping product name processing due to missing column", fg=typer.colors.YELLOW))
        else:
            typer.echo(typer.style(
                f"Found '{config.product_name_column}' column at index {product_name_col}", fg=typer.colors.GREEN
            ))

            # Step 7: Filter rows based on keywords (FIRST)
            typer.echo(typer.style("Step 7: Filtering rows with keywords", fg=typer.colors.BLUE, bold=True))
            cleared_count = clear_filtered_rows(
                new_sheet, product_name_col, config.filter_keywords
            )
            typer.echo(typer.style(f"Row filtering completed: {cleared_count} rows cleared", fg=typer.colors.YELLOW))

            # Step 8: Clean product names by removing content between 【】 (SECOND)
            typer.echo(typer.style("Step 8: Cleaning product names by removing 【】 content", fg=typer.colors.BLUE, bold=True))
            cleaned_count = clean_product_names(new_sheet, product_name_col)
            typer.echo(typer.style(
                f"Product name cleaning completed: {cleaned_count} names cleaned", fg=typer.colors.YELLOW
            ))

            # Step 9: Process product names with Dify AI
            typer.echo(typer.style("Step 9: Processing product names with Dify AI", fg=typer.colors.BLUE, bold=True))
            try:
                processor = DifyProductProcessor(
                    base_url="http://119.45.130.88:8080/v1",
                    api_key="app-n2xjLW2SqEDSuq2RA7O6oqwW",  # 请替换为实际的 API Key
                    user_id="process_goods_user"
                )
                
                # 提取唯一商品名称
                unique_products = processor.extract_unique_product_names(new_sheet, product_name_col)
                
                if unique_products:
                    # 处理所有商品
                    processed_results = processor.process_all_products(unique_products)
                    
                    # 更新工作表中的商品名称
                    processor.update_worksheet_with_processed_names(
                        new_sheet, product_name_col, unique_products, processed_results
                    )
                    
                    typer.echo(f"Dify AI processing completed for {len(processed_results)} products")
                else:
                    typer.echo("No products found for Dify processing")
                    
            except Exception as e:
                typer.echo(f"Dify processing failed, continuing without AI processing: {e}", err=True)

        # Final step: Save the workbook
        typer.echo("Final step: Saving processed workbook")
        workbook.save(config.output_path)
        output_size = config.output_path.stat().st_size
        typer.echo("Processing completed successfully!")
        typer.echo(f"Output saved to: {config.output_path} ({output_size} bytes)")
        typer.echo(typer.style("=== Processing finished ===", fg=typer.colors.MAGENTA, bold=True))

    except Exception as e:
        typer.echo(typer.style(f"Error during processing: {e}", fg=typer.colors.RED, bold=True), err=True)
        sys.exit(1)


@app.command()
def main(
    file_path: Path = typer.Argument(..., help="Path to the Excel file to process"),
    output_path: Optional[Path] = typer.Option(
        None, "--output", "-o", help="Output file path (optional)"
    ),
    filter_keywords: Optional[str] = typer.Option(
        "员工, 测试", "--filter", "-f", help="Comma-separated keywords to filter"
    ),
    product_column: str = typer.Option(
        "商品名称",
        "--column",
        "-c",
        help="Product name column name (default: 商品名称)",
    ),
):
    """
    Process goods.xlsx file by copying it and creating a filtered duplicate sheet.

    This script will:
    1. Copy the input file to a new file
    2. Open the new file
    3. Create a new sheet named 'sheet1' as a copy of the first sheet
    4. Filter out rows containing specified keywords in the product name column
    5. Clean product names by removing content between 【】 characters
    """
    try:
        typer.echo(typer.style("Starting goods processing script", fg=typer.colors.MAGENTA, bold=True))
        typer.echo(typer.style(f"Input file: {file_path}", fg=typer.colors.CYAN))
        typer.echo(typer.style(f"Output file: {output_path or 'auto-generated'}", fg=typer.colors.CYAN))
        typer.echo(typer.style(f"Filter keywords: {filter_keywords}", fg=typer.colors.CYAN))
        typer.echo(typer.style(f"Product column: {product_column}", fg=typer.colors.CYAN))

        # Parse filter keywords
        keywords_list = [k.strip() for k in filter_keywords.split(",") if k.strip()]
        typer.echo(typer.style(f"Parsed filter keywords: {keywords_list}", fg=typer.colors.GREEN))

        config = FileConfig(
            input_path=file_path,
            output_path=output_path,
            filter_keywords=keywords_list,
            product_name_column=product_column,
        )
        typer.echo(typer.style("Configuration created successfully", fg=typer.colors.GREEN))

        process_excel_file(config)

    except Exception as e:
        typer.echo(typer.style(f"Configuration error: {e}", fg=typer.colors.RED, bold=True), err=True)
        sys.exit(1)


if __name__ == "__main__":
    app()
